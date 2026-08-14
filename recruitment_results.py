import re
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZipFile


VALID_RESULT_STATUSES = ("pending", "accepted", "waitlisted", "rejected")
RESULT_STATUS_LABELS = {
    "pending": "발표 전",
    "accepted": "합격",
    "waitlisted": "예비",
    "rejected": "불합격",
}
RESULT_STATUS_ALIASES = {
    "": "pending",
    "pending": "pending",
    "발표전": "pending",
    "미정": "pending",
    "accepted": "accepted",
    "합격": "accepted",
    "waitlisted": "waitlisted",
    "waitlist": "waitlisted",
    "예비": "waitlisted",
    "예비합격": "waitlisted",
    "rejected": "rejected",
    "불합격": "rejected",
}


def normalize_applicant_name(value):
    """Match names case-insensitively while ignoring all whitespace."""
    return re.sub(r"\s+", "", str(value or "")).casefold()


def normalize_student_id(value):
    return re.sub(r"\s+", "", str(value or ""))


def normalize_result_status(value):
    key = re.sub(r"\s+", "", str(value or "")).casefold()
    return RESULT_STATUS_ALIASES.get(key)


def parse_applicant_rows(raw_text, max_rows=500):
    """Parse rows pasted from Excel: name, student id, optional status/message."""
    parsed = []
    errors = []
    seen_student_ids = set()
    lines = str(raw_text or "").splitlines()
    non_empty_lines = [line for line in lines if line.strip()]
    if len(non_empty_lines) > max_rows:
        return [], [f"한 번에 최대 {max_rows}명까지 입력할 수 있습니다."]

    for line_number, line in enumerate(lines, 1):
        if not line.strip():
            continue
        parts = line.split("\t") if "\t" in line else line.split(",")
        parts = [part.strip() for part in parts]
        if len(parts) < 2:
            errors.append(f"{line_number}행: 이름과 학번이 필요합니다.")
            continue
        if not parsed and parts[0].replace(" ", "") in ("이름", "성명") and parts[1].replace(" ", "") in ("학번", "학생번호"):
            continue
        name = parts[0]
        student_id = normalize_student_id(parts[1])
        status_provided = len(parts) > 2 and bool(parts[2].strip())
        message_provided = len(parts) > 3
        status = normalize_result_status(parts[2] if status_provided else "")
        personal_message = ",".join(parts[3:]).strip() if "\t" not in line else "\t".join(parts[3:]).strip()
        duplicate_student_id = student_id in seen_student_ids
        if re.fullmatch(r"[0-9]{4,20}", student_id):
            seen_student_ids.add(student_id)
        if not normalize_applicant_name(name):
            errors.append(f"{line_number}행: 이름을 확인해주세요.")
        elif not re.fullmatch(r"[0-9]{4,20}", student_id):
            errors.append(f"{line_number}행: 학번은 4~20자리 숫자여야 합니다.")
        elif duplicate_student_id:
            errors.append(f"{line_number}행: 같은 학번이 입력 안에 중복되어 있습니다.")
        elif status is None:
            errors.append(f"{line_number}행: 결과는 발표 전·합격·예비·불합격 중 하나여야 합니다.")
        elif len(personal_message) > 3000:
            errors.append(f"{line_number}행: 개인 안내는 3,000자 이하여야 합니다.")
        else:
            parsed.append({
                "name": name,
                "name_key": normalize_applicant_name(name),
                "student_id": student_id,
                "result_status": status,
                "personal_message": personal_message or None,
                "_status_provided": status_provided,
                "_message_provided": message_provided,
            })
    return parsed, errors


def _excel_cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_applicant_file(filename, content, max_rows=500, max_bytes=2 * 1024 * 1024):
    """Parse a small .xlsx or .csv applicant list without writing it to disk."""
    suffix = Path(str(filename or "")).suffix.lower()
    raw = bytes(content or b"")
    if not raw:
        return [], ["업로드한 파일이 비어 있습니다."]
    if len(raw) > max_bytes:
        return [], ["명단 파일은 2MB 이하만 업로드할 수 있습니다."]
    if suffix == ".csv":
        try:
            text = raw.decode("utf-8-sig")
        except UnicodeDecodeError:
            try:
                text = raw.decode("cp949")
            except UnicodeDecodeError:
                return [], ["CSV 파일은 UTF-8 또는 한글 Excel 형식으로 저장해주세요."]
        return parse_applicant_rows(text, max_rows=max_rows)
    if suffix != ".xlsx":
        return [], ["Excel(.xlsx) 또는 CSV(.csv) 파일만 업로드할 수 있습니다."]

    try:
        with ZipFile(BytesIO(raw)) as archive:
            if sum(item.file_size for item in archive.infolist()) > 20 * 1024 * 1024:
                return [], ["압축을 푼 명단 파일이 너무 큽니다."]
        from openpyxl import load_workbook
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        lines = []
        for values in sheet.iter_rows(values_only=True):
            cells = [_excel_cell_text(value) for value in values[:4]]
            if any(cells):
                lines.append("\t".join(cells))
                if len(lines) > max_rows + 1:
                    workbook.close()
                    return [], [f"한 번에 최대 {max_rows}명까지 입력할 수 있습니다."]
        workbook.close()
    except (BadZipFile, OSError, ValueError):
        return [], ["Excel 파일을 읽을 수 없습니다. 손상되지 않은 .xlsx 파일인지 확인해주세요."]
    except Exception:
        return [], ["Excel 파일 형식을 확인해주세요."]
    return parse_applicant_rows("\n".join(lines), max_rows=max_rows)
