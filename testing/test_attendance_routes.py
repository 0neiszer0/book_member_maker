import importlib
import os
from io import BytesIO
from types import SimpleNamespace
import unittest
from unittest.mock import patch
from testing._fake_supabase import FakeSupabase

os.environ.setdefault('PYTHON_DOTENV_DISABLED', '1')
os.environ.setdefault('FLASK_SECRET_KEY', 'attendance-route-test')
os.environ.setdefault('SUPABASE_URL', 'http://127.0.0.1:9')
os.environ.setdefault('SUPABASE_SERVICE_KEY', 'attendance-route-test')


class AttendanceRouteTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.module = importlib.import_module('app')
        cls.module.app.config['TESTING'] = True

    def setUp(self):
        self.original = self.module.supabase
        self.fake = FakeSupabase({'members': [
            {'id': 1, 'name': '운영', 'student_id': '202600001', 'role': 'admin', 'is_active': True, 'member_status': 'active', 'account_status': 'active'},
            {'id': 2, 'name': '호반', 'student_id': '202600002', 'is_active': True}],
            'seminar_terms': [{'id': 'term', 'name': '학기', 'start_date': '2026-08-01', 'end_date': '2026-08-31', 'is_active': True}],
            'seminar_sessions': [{'id': 'thu', 'term_id': 'term', 'day_type': 'thu', 'meeting_date': '2026-08-13', 'planned_member_ids': None},
                                 {'id': 'mon', 'term_id': 'term', 'day_type': 'mon', 'meeting_date': '2026-08-17', 'planned_member_ids': []}]})
        self.module.supabase = self.fake
        self.client = self.module.app.test_client()
        with self.client.session_transaction() as state:
            state.update(user_id=1, user_role='admin', user_name='운영', attendance_csrf='csrf')
        self.headers = {'X-CSRF-Token': 'csrf'}

    def tearDown(self):
        self.module.supabase = self.original

    def test_new_pages_render_and_self_page_has_no_other_member(self):
        for path in ['/admin/seminar_sessions/thu/roster', '/admin/term_attendance?term_id=term', '/my/term_attendance?term_id=term']:
            with self.subTest(path=path):
                response = self.client.get(path)
                self.assertEqual(response.status_code, 200)
        text = self.client.get('/my/term_attendance?term_id=term').get_data(as_text=True)
        self.assertNotIn('202600002', text)

    def test_preview_rejects_csrf(self):
        response = self.client.post('/api/admin/seminar_sessions/thu/roster/preview', json={'method': 'selected', 'member_ids': [2]})
        self.assertEqual(response.status_code, 403)

    def test_preview_token_applies_through_atomic_rpc(self):
        preview = self.client.post('/api/admin/seminar_sessions/thu/roster/preview', headers=self.headers, json={'method': 'text', 'text': '호반 202600002'}).get_json()
        captured = {}
        def rpc(name, params):
            captured.update(name=name, params=params)
            return SimpleNamespace(execute=lambda: SimpleNamespace(data={'accepted': True}))
        self.fake.rpc = rpc
        response = self.client.post('/api/admin/seminar_sessions/thu/roster/apply', headers=self.headers, json={'token': preview['token']})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured['name'], 'apply_seminar_roster')
        self.assertEqual(captured['params']['p_member_ids'], [2])
        self.assertEqual(captured['params']['p_expected_member_ids'], [1])

    def test_unknown_names_prevent_partial_apply_token(self):
        result = self.client.post('/api/admin/seminar_sessions/mon/roster/preview', headers=self.headers, json={'method': 'text', 'text': '운영\n모르는사람'}).get_json()
        self.assertIsNone(result['token'])
        self.assertEqual(len(result['issues']), 1)

    def test_xlsx_preview_and_export(self):
        from openpyxl import Workbook, load_workbook
        wb = Workbook(); wb.active.append(['이름', '학번']); wb.active.append(['호반', '202600002'])
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        response = self.client.post('/api/admin/seminar_sessions/mon/roster/preview', headers=self.headers,
                                    data={'method': 'file', 'file': (buf, 'roster.xlsx')})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.get_json()['expected_count'], 1)
        exported = self.client.get('/admin/term_attendance/export?term_id=term')
        self.assertEqual(exported.status_code, 200)
        sheet = load_workbook(BytesIO(exported.data)).active
        self.assertEqual(sheet.cell(2, 8).value, 0)

    def test_conflicting_rpc_does_not_claim_success(self):
        preview = self.client.post('/api/admin/seminar_sessions/thu/roster/preview', headers=self.headers, json={'method': 'selected', 'member_ids': []}).get_json()
        self.fake.rpc = lambda *args: SimpleNamespace(execute=lambda: SimpleNamespace(data={'accepted': False}))
        result = self.client.post('/api/admin/seminar_sessions/thu/roster/apply', headers=self.headers, json={'token': preview['token']})
        self.assertEqual(result.status_code, 409)

    def test_legacy_signup_routes_are_retired_without_data_reads(self):
        for path in ['/api/seminar_vote/verify', '/api/seminar_vote/submit', '/api/attendance']:
            result = self.client.post(path, json={})
            self.assertEqual(result.status_code, 410)
        self.assertEqual(self.client.get('/api/seminar_vote/counts').status_code, 410)
        self.assertFalse(any(call[0] in {'update', 'insert', 'delete'} for call in self.fake.calls))


if __name__ == '__main__':
    unittest.main()
