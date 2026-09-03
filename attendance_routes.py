"""Admin-only roster imports and term participation reporting."""
from datetime import datetime, timedelta, timezone
from io import BytesIO
import secrets
from flask import abort, jsonify, render_template, request, session, send_file
from itsdangerous import URLSafeTimedSerializer, BadSignature, SignatureExpired
from attendance_workflow import (build_term_attendance, expected_member_ids,
                                 match_roster_rows, parse_roster_text)

KST = timezone(timedelta(hours=9))


def init_attendance_routes(app, get_supabase, login_required, rebuild_matrix=None):
    signer = URLSafeTimedSerializer(app.secret_key, salt='seminar-roster-preview-v1')

    def db():
        return get_supabase()

    def csrf_token():
        if not session.get('attendance_csrf'):
            session['attendance_csrf'] = secrets.token_urlsafe(32)
        return session['attendance_csrf']

    def check_csrf():
        supplied = request.headers.get('X-CSRF-Token') or request.form.get('csrf_token') or ''
        expected = session.get('attendance_csrf') or ''
        if not expected or not secrets.compare_digest(supplied, expected):
            abort(403)

    def roster_context(session_id):
        target = db().table('seminar_sessions').select('*').eq('id', session_id).single().execute().data
        if not target:
            abort(404)
        if target.get('day_type') not in {'mon', 'thu'}:
            abort(400)
        members = db().table('members').select('id,name,student_id,department,is_active').order('name').execute().data or []
        audit = db().table('seminar_roster_imports').select('*').eq('session_id', session_id).order('created_at', desc=True).limit(10).execute().data or []
        mode = 'absence' if target.get('day_type') == 'thu' else 'attendance'
        active = {int(m['id']) for m in members if m.get('is_active')}
        if mode == 'absence' and audit:
            eligible = set(audit[0].get('member_ids') or []) | set(audit[0].get('expected_member_ids') or [])
        else:
            eligible = active | set(target.get('planned_member_ids') or [])
        if target.get('planned_member_ids') is not None:
            planned = set(target['planned_member_ids'])
        elif mode == 'absence':
            absent = db().table('seminar_absences').select('member_id').eq('session_id', session_id).is_('cancelled_at', 'null').execute().data or []
            planned = eligible - {r['member_id'] for r in absent}
        else:
            votes = db().table('seminar_votes').select('member_id').eq('session_id', session_id).eq('attending', True).execute().data or []
            planned = {r['member_id'] for r in votes}
        no_shows = db().table('seminar_no_shows').select('member_id').eq('session_id', session_id).is_('cancelled_at', 'null').execute().data or []
        return target, members, audit, mode, sorted(eligible), sorted(planned), {r['member_id'] for r in no_shows}

    @app.get('/admin/seminar_sessions/<session_id>/roster')
    @login_required(role='admin')
    def admin_seminar_roster(session_id):
        target, members, audit, mode, eligible, planned, no_shows = roster_context(session_id)
        input_ids = sorted(set(eligible) - set(planned)) if mode == 'absence' else planned
        actual = target.get('actual_member_ids')
        if actual is None:
            actual = sorted(set(planned) - no_shows)
            histories = db().table('history').select('present,actual_member_ids,attendance_confirmed_at') \
                .eq('seminar_session_id', session_id).execute().data or []
            confirmed = [h for h in histories if h.get('attendance_confirmed_at')]
            if confirmed:
                known_ids = set()
                for history in confirmed:
                    if history.get('actual_member_ids') is not None:
                        known_ids.update(history['actual_member_ids'])
                    else:
                        for name in history.get('present') or []:
                            matches = [m['id'] for m in members if m.get('name') == name]
                            if len(matches) == 1:
                                known_ids.add(matches[0])
                actual = sorted(known_ids - no_shows)
        def local_stamp(value):
            if not value:
                return '아직 반영하지 않음'
            return datetime.fromisoformat(value.replace('Z', '+00:00')).astimezone(KST).strftime('%Y-%m-%d %H:%M')
        target['roster_updated_label'] = local_stamp(target.get('roster_updated_at'))
        for row in audit:
            row['created_label'] = local_stamp(row.get('created_at'))
        return render_template('admin_seminar_roster.html', seminar=target, members=members,
                               audit=audit, mode=mode, eligible_ids=eligible, input_ids=input_ids,
                               planned_ids=planned, actual_ids=actual, csrf_token=csrf_token(),
                               today=datetime.now(KST).date().isoformat())

    @app.post('/api/admin/seminar_sessions/<session_id>/roster/preview')
    @login_required(role='admin')
    def seminar_roster_preview(session_id):
        check_csrf()
        try:
            target, members, audit, mode, eligible, planned, _ = roster_context(session_id)
            body = request.get_json(silent=True) or request.form
            method = body.get('method', 'text')
            if method not in {'text', 'file', 'selected'}:
                raise ValueError('명단 입력 방식을 다시 선택해주세요.')
            selectable = [m for m in members if int(m['id']) in eligible]
            issues = []
            if method == 'selected':
                raw_ids = body.get('member_ids', [])
                if not isinstance(raw_ids, list) or len(raw_ids) > 2000:
                    raise ValueError('선택 명단이 올바르지 않습니다.')
                ids = [int(mid) for mid in raw_ids]
                if len(set(ids)) != len(ids):
                    raise ValueError('같은 회원이 두 번 선택되어 있습니다.')
            else:
                if method == 'file':
                    upload = request.files.get('file')
                    if not upload or not upload.filename.lower().endswith('.xlsx'):
                        raise ValueError('이름·학번 열이 있는 .xlsx 파일을 선택해주세요.')
                    content = upload.read(2 * 1024 * 1024 + 1)
                    if len(content) > 2 * 1024 * 1024:
                        raise ValueError('엑셀 파일은 2MB 이하로 올려주세요.')
                    from zipfile import ZipFile
                    with ZipFile(BytesIO(content)) as archive:
                        if sum(info.file_size for info in archive.infolist()) > 20 * 1024 * 1024:
                            raise ValueError('압축을 푼 엑셀 파일이 너무 큽니다.')
                    from openpyxl import load_workbook
                    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
                    sheet = workbook.active
                    if sheet.max_row > 2001 or sheet.max_column > 10:
                        raise ValueError('엑셀은 2,000명, 10열 이하로 정리해주세요.')
                    raw_rows = [[str(v).strip() if v is not None else '' for v in row] for row in sheet.iter_rows(values_only=True)]
                    workbook.close()
                    if raw_rows:
                        headers = [value.lower() for value in raw_rows[0]]
                        name_col = next((i for i, value in enumerate(headers) if value in {'이름', '성명', 'name'}), None)
                        id_col = next((i for i, value in enumerate(headers) if value in {'학번', 'student_id'}), None)
                        if name_col is not None or id_col is not None:
                            columns = [i for i in (name_col, id_col) if i is not None]
                            raw_rows = [[row[i] if i < len(row) else '' for i in columns] for row in raw_rows[1:]]
                else:
                    raw_rows = parse_roster_text(str(body.get('text') or ''))
                matches, issues = match_roster_rows(raw_rows, selectable)
                ids = [int(m['id']) for m in matches]
            expected = expected_member_ids(mode, ids, eligible)
            by_id = {int(m['id']): m for m in members}
            payload = {'session_id': session_id, 'user_id': session.get('user_id'), 'mode': mode,
                       'member_ids': sorted(ids), 'expected_member_ids': expected,
                       'previous_member_ids': planned, 'previous_updated_at': target.get('roster_updated_at')}
            return jsonify(status='success', issues=issues,
                           matched=[by_id[mid] for mid in sorted(ids)],
                           added=[by_id[mid] for mid in sorted(set(expected) - set(planned))],
                           removed=[by_id[mid] for mid in sorted(set(planned) - set(expected))],
                           expected_count=len(expected), token=signer.dumps(payload) if not issues else None)
        except (ValueError, TypeError) as exc:
            return jsonify(status='error', message=str(exc)), 400
        except Exception:
            app.logger.exception('Roster preview failed')
            return jsonify(status='error', message='명단을 읽지 못했습니다. 파일 형식과 회원 정보를 확인해주세요.'), 400

    def apply_preview(payload):
        try:
            result = db().rpc('apply_seminar_roster', {
                'p_session_id': payload['session_id'], 'p_member_ids': payload['member_ids'],
                'p_expected_member_ids': payload['expected_member_ids'],
                'p_previous_member_ids': payload['previous_member_ids'],
                'p_previous_updated_at': payload['previous_updated_at'], 'p_mode': payload['mode'],
                'p_created_by': session.get('user_id'),
            }).execute().data or {}
        except Exception:
            app.logger.exception('Atomic roster apply failed')
            return jsonify(status='error', message='명단 반영 결과를 확인하지 못했습니다. 새로고침하여 반영 내역을 먼저 확인해주세요.'), 503
        if not result.get('accepted'):
            return jsonify(status='error', message='다른 변경이 반영되었습니다. 새로고침 후 명단을 다시 확인해주세요.'), 409
        return jsonify(status='success')

    @app.post('/api/admin/seminar_sessions/<session_id>/roster/apply')
    @login_required(role='admin')
    def seminar_roster_apply(session_id):
        check_csrf()
        try:
            payload = signer.loads((request.get_json(silent=True) or {}).get('token', ''), max_age=900)
            if payload.get('session_id') != session_id or payload.get('user_id') != session.get('user_id'):
                abort(403)
            return apply_preview(payload)
        except (BadSignature, SignatureExpired):
            return jsonify(status='error', message='미리보기가 만료되었습니다. 다시 확인해주세요.'), 400

    @app.post('/api/admin/seminar_sessions/<session_id>/roster/undo')
    @login_required(role='admin')
    def seminar_roster_undo(session_id):
        check_csrf()
        target, members, audit, mode, eligible, planned, _ = roster_context(session_id)
        if not audit:
            return jsonify(status='error', message='되돌릴 명단 반영 내역이 없습니다.'), 400
        previous = sorted(audit[0].get('previous_member_ids') or [])
        return apply_preview({'session_id': session_id, 'mode': mode,
                              'member_ids': sorted(set(eligible) - set(previous)) if mode == 'absence' else previous,
                              'expected_member_ids': previous, 'previous_member_ids': planned,
                              'previous_updated_at': target.get('roster_updated_at')})

    @app.post('/api/admin/seminar_sessions/<session_id>/attendance/confirm')
    @login_required(role='admin')
    def seminar_attendance_confirm(session_id):
        check_csrf()
        target, members, _, _, _, _, no_shows = roster_context(session_id)
        if target.get('meeting_date', '') > datetime.now(KST).date().isoformat():
            return jsonify(status='error', message='아직 진행하지 않은 세미나는 출석을 확정할 수 없습니다.'), 400
        body = request.get_json(silent=True) or {}
        try:
            if not isinstance(body.get('member_ids'), list):
                raise ValueError('실제 참석 명단을 선택해주세요.')
            ids = sorted({int(mid) for mid in body['member_ids']})
            if set(ids) - {int(m['id']) for m in members}:
                raise ValueError('존재하지 않는 회원이 포함되어 있습니다.')
            if set(ids) & no_shows:
                raise ValueError('미연락 불참자로 기록된 회원이 있습니다. 먼저 불참 기록을 취소해주세요.')
        except (ValueError, TypeError) as exc:
            return jsonify(status='error', message=str(exc)), 400
        stamp = datetime.now(timezone.utc).isoformat()
        db().table('seminar_sessions').update({'actual_member_ids': ids, 'attendance_confirmed_at': stamp}).eq('id', session_id).execute()
        # History remains an editable group plan; confirmation is separate metadata.
        db().table('history').update({'attendance_confirmed_at': stamp}).eq('seminar_session_id', session_id).execute()
        if rebuild_matrix:
            rebuild_matrix(session_id)
        return jsonify(status='success')

    def term_report(term_id=None):
        terms = db().table('seminar_terms').select('*').order('start_date', desc=True).execute().data or []
        term = next((t for t in terms if str(t['id']) == str(term_id)), None)
        if not term:
            today = datetime.now(KST).date().isoformat()
            term = next((t for t in terms if t['start_date'] <= today <= t['end_date']), None)
        if not term:
            term = next((t for t in terms if t.get('is_active')), terms[0] if terms else None)
        if not term:
            return terms, None, None
        start, end = term['start_date'], term['end_date']
        members = db().table('members').select('id,name,student_id,department,is_active').order('name').execute().data or []
        histories = db().table('history').select('id,date,present,book_title,seminar_session_id,attendance_confirmed_at,actual_member_ids').gte('date', start).lte('date', end).execute().data or []
        sessions = db().table('seminar_sessions').select('*').gte('meeting_date', start).lte('meeting_date', end).execute().data or []
        session_ids = [s['id'] for s in sessions]
        no_shows = db().table('seminar_no_shows').select('session_id,member_id,cancelled_at').in_('session_id', session_ids).is_('cancelled_at', 'null').execute().data or [] if session_ids else []
        events = db().table('special_events').select('*').gte('event_date', start).lte('event_date', end).execute().data or []
        event_ids = [e['id'] for e in events]
        attendees = db().table('special_event_attendees').select('event_id,member_id').in_('event_id', event_ids).execute().data or [] if event_ids else []
        bricks = db().table('brick_book_sessions').select('id,meeting_date,brick_books(title)').gte('meeting_date', start).lte('meeting_date', end).execute().data or []
        brick_ids = [b['id'] for b in bricks]
        brick_members = db().table('brick_session_members').select('session_id,member_id').in_('session_id', brick_ids).execute().data or [] if brick_ids else []
        report = build_term_attendance(members, histories, sessions, events, attendees, bricks, brick_members,
                                       start, end, term.get('attendance_minimum', 3), datetime.now(KST).date(), no_shows)
        report['events'] = events
        report['pending_sessions'] = [s for s in sessions if s['meeting_date'] <= datetime.now(KST).date().isoformat()
                                      and not s.get('attendance_confirmed_at')
                                      and not any(h.get('attendance_confirmed_at') and h.get('seminar_session_id') == s['id'] for h in histories)]
        return terms, term, report

    @app.get('/admin/term_attendance')
    @login_required(role='admin')
    def admin_term_attendance():
        terms, term, report = term_report(request.args.get('term_id'))
        return render_template('admin_term_attendance.html', terms=terms, term=term, report=report,
                               csrf_token=csrf_token(), show=request.args.get('show', 'all'))

    @app.post('/api/admin/term_attendance/<term_id>/minimum')
    @login_required(role='admin')
    def term_attendance_minimum(term_id):
        check_csrf()
        try:
            value = int((request.get_json(silent=True) or {}).get('minimum'))
            if not 1 <= value <= 50:
                raise ValueError()
        except (ValueError, TypeError):
            return jsonify(status='error', message='최소 참석 기준은 1~50회로 입력해주세요.'), 400
        db().table('seminar_terms').update({'attendance_minimum': value}).eq('id', term_id).execute()
        return jsonify(status='success')

    @app.post('/api/admin/term_attendance/events/<event_id>/confirm')
    @login_required(role='admin')
    def term_attendance_event_confirm(event_id):
        check_csrf()
        event = db().table('special_events').select('*').eq('id', event_id).single().execute().data or {}
        if not event or event.get('event_date', '') > datetime.now(KST).date().isoformat():
            return jsonify(status='error', message='진행이 끝난 OT의 실제 참석 명단을 먼저 확인해주세요.'), 400
        enabled = (request.get_json(silent=True) or {}).get('enabled') is True
        db().table('special_events').update({'counts_toward_attendance': enabled,
            'attendance_confirmed_at': datetime.now(timezone.utc).isoformat() if enabled else None}).eq('id', event_id).execute()
        return jsonify(status='success')

    @app.get('/admin/term_attendance/export')
    @login_required(role='admin')
    def term_attendance_export():
        _, term, report = term_report(request.args.get('term_id'))
        if not report:
            abort(404)
        from openpyxl import Workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = '학기 참석 현황'
        sheet.append(['이름', '학번', '소속', '상태', '세미나(주차)', 'OT', '벽돌책', '합계', '부족 횟수'] +
                     [f"{c['date']} {c['title']}" for c in report['columns']])
        def safe(value):
            return "'" + value if isinstance(value, str) and value.startswith(('=', '+', '-', '@')) else value
        for member in report['members']:
            row = [member.get('name'), member.get('student_id'), member.get('department'),
                   '활성' if member.get('is_active') else '비활성·휴면',
                   member['counts']['seminar'], member['counts']['ot'], member['counts']['brick'],
                   member['total'], member['shortage']] + [
                       'O' if report['matrix'].get(member['id'], {}).get(c['key']) else '' for c in report['columns']]
            sheet.append([safe(v) for v in row])
        sheet.freeze_panes = 'E2'
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 20
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True, download_name='term_attendance.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    @app.get('/my/term_attendance')
    @login_required(role='ANY')
    def my_term_attendance():
        terms, term, report = term_report(request.args.get('term_id'))
        mine = next((m for m in (report or {}).get('members', []) if int(m['id']) == int(session['user_id'])), None)
        columns = [c for c in (report or {}).get('columns', []) if mine and report['matrix'].get(mine['id'], {}).get(c['key'])]
        return render_template('my_term_attendance.html', term=term, mine=mine, columns=columns,
                               minimum=(report or {}).get('minimum', 3))
